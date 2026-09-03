#!/usr/bin/env python3
"""CV Screener — batch-screen PDF resumes against your criteria into an Excel sheet.

Usage:
    python3 cv_screener.py --pdf-dir ./resumes --criteria criteria.json --out results.xlsx
    python3 cv_screener.py --pdf-dir ./resumes --criteria criteria.json \
        --template your_sheet.xlsx --out results.xlsx

What it does, per PDF:
  1. Extracts text (pdfplumber, falls back to pypdf).
  2. If the PDF has no text layer (scanned/image-based), runs OCR when pytesseract +
     tesseract are installed; otherwise the row is flagged NEEDS_OCR — never silently skipped.
  3. Pulls candidate fields: name (heuristic), email, phone, years of experience, education level.
  4. Scores the CV against your criteria (skills with weights, required skills, minimum years).
  5. Writes one row per CV into the output Excel. If you pass --template, it appends rows under
     your existing headers (matching by header name, case-insensitive).

Criteria file (criteria.json) example:
{
  "skills":   {"python": 10, "sql": 6, "aws": 5, "excel": 3},
  "required": ["python"],
  "min_years": 2,
  "education_bonus": {"phd": 5, "master": 3, "bachelor": 2}
}

Dependencies: pip install pdfplumber pypdf openpyxl   (optional OCR: pip install pytesseract pillow
+ install the tesseract binary). Pure local processing — no resume data leaves your machine.
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path

EMAIL_RE = re.compile(r"[\w.+-]+@[\w-]+\.[\w.-]+")
PHONE_RE = re.compile(r"(?:\+?\d{1,3}[\s.-]?)?(?:\(?\d{3}\)?[\s.-]?)\d{3}[\s.-]?\d{4}")
YEARS_RE = re.compile(r"(\d{1,2})\s*\+?\s*years?(?:\s+of)?\s+(?:\w+\s+){0,2}experience", re.I)
DEGREES = [("phd", r"\bph\.?d\b|\bdoctorate\b"), ("master", r"\bm\.?sc?\.?\b|\bmaster'?s?\b|\bmba\b"),
           ("bachelor", r"\bb\.?sc?\.?\b|\bbachelor'?s?\b|\bb\.?a\.?\b|\bb\.?eng\b")]
MIN_TEXT_CHARS = 80   # below this we assume a scanned/image PDF


def extract_text(pdf_path: Path) -> tuple[str, str]:
    """Return (text, method). Tries pdfplumber, then pypdf, then OCR if available."""
    text = ""
    try:
        import pdfplumber
        with pdfplumber.open(pdf_path) as pdf:
            text = "\n".join(page.extract_text() or "" for page in pdf.pages)
        if len(text.strip()) >= MIN_TEXT_CHARS:
            return text, "text"
    except Exception:
        pass
    try:
        from pypdf import PdfReader
        text = "\n".join((p.extract_text() or "") for p in PdfReader(str(pdf_path)).pages)
        if len(text.strip()) >= MIN_TEXT_CHARS:
            return text, "text"
    except Exception:
        pass
    # scanned/image PDF -> OCR if the toolchain exists
    try:
        import pytesseract  # noqa: F401
        from pdf2image import convert_from_path
        pages = convert_from_path(str(pdf_path))
        ocr = "\n".join(pytesseract.image_to_string(img) for img in pages)
        if ocr.strip():
            return ocr, "ocr"
    except Exception:
        pass
    return text, "needs_ocr"


def parse_fields(text: str) -> dict:
    email = EMAIL_RE.search(text)
    phone = PHONE_RE.search(text)
    years = 0
    for m in YEARS_RE.finditer(text):
        years = max(years, int(m.group(1)))
    education = ""
    for level, pat in DEGREES:
        if re.search(pat, text, re.I):
            education = level
            break
    # name heuristic: first short, letters-only line that isn't a heading
    name = ""
    for line in (l.strip() for l in text.splitlines()):
        if 4 <= len(line) <= 40 and re.fullmatch(r"[A-Za-z][A-Za-z .,'-]+", line) \
                and line.lower() not in ("curriculum vitae", "resume", "cv"):
            name = line
            break
    return {"name": name, "email": email.group(0) if email else "",
            "phone": phone.group(0) if phone else "", "years_experience": years,
            "education": education}


def score(text: str, fields: dict, criteria: dict) -> tuple[float, list, list]:
    low = text.lower()
    matched, missing_required = [], []
    total = 0.0
    for skill, weight in (criteria.get("skills") or {}).items():
        if skill.lower() in low:
            matched.append(skill)
            total += float(weight)
    for req in (criteria.get("required") or []):
        if req.lower() not in low:
            missing_required.append(req)
    min_years = int(criteria.get("min_years") or 0)
    if min_years and fields["years_experience"] >= min_years:
        total += 5
    bonus = (criteria.get("education_bonus") or {}).get(fields["education"], 0)
    total += float(bonus)
    if missing_required:
        total *= 0.5   # missing a hard requirement halves the score (still visible, not hidden)
    return round(total, 1), matched, missing_required


COLUMNS = ["File", "Name", "Email", "Phone", "Years Experience", "Education",
           "Skills Matched", "Missing Required", "Score", "Extraction", "Notes"]


def main() -> int:
    ap = argparse.ArgumentParser(description="Screen PDF CVs into an Excel sheet.")
    ap.add_argument("--pdf-dir", required=True)
    ap.add_argument("--criteria", required=True)
    ap.add_argument("--template", default=None, help="existing .xlsx whose headers to fill")
    ap.add_argument("--out", default="screening_results.xlsx")
    args = ap.parse_args()

    pdf_dir = Path(args.pdf_dir)
    pdfs = sorted(pdf_dir.glob("*.pdf"))
    if not pdfs:
        print(f"no PDFs found in {pdf_dir}"); return 1
    criteria = json.loads(Path(args.criteria).read_text())

    from openpyxl import Workbook, load_workbook
    if args.template:
        wb = load_workbook(args.template)
        ws = wb.active
        headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
    else:
        wb = Workbook(); ws = wb.active; ws.title = "Screening"
        ws.append(COLUMNS); headers = COLUMNS

    hmap = {h.lower(): i for i, h in enumerate(headers)}
    done = needs_ocr = 0
    for pdf in pdfs:
        try:
            text, method = extract_text(pdf)
            if method == "needs_ocr":
                row_data = {"file": pdf.name, "extraction": "NEEDS_OCR",
                            "notes": "scanned PDF; install tesseract+pytesseract+pdf2image or scan manually"}
                needs_ocr += 1
            else:
                fields = parse_fields(text)
                s, matched, missing = score(text, fields, criteria)
                row_data = {"file": pdf.name, "name": fields["name"], "email": fields["email"],
                            "phone": fields["phone"], "years experience": fields["years_experience"],
                            "education": fields["education"], "skills matched": ", ".join(matched),
                            "missing required": ", ".join(missing), "score": s,
                            "extraction": method, "notes": ""}
                done += 1
            row = [""] * len(headers)
            for key, val in row_data.items():
                idx = hmap.get(key)
                if idx is not None:
                    row[idx] = val
            ws.append(row)
            print(f"  {pdf.name}: {row_data.get('extraction')} score={row_data.get('score','-')}")
        except Exception as e:                      # one bad file never kills the batch
            ws.append([pdf.name] + [""] * (len(headers) - 2) + [f"ERROR: {e}"])
            print(f"  {pdf.name}: ERROR {e}")

    # sort by score (descending) when we created the sheet ourselves
    if not args.template:
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        score_i = COLUMNS.index("Score")
        rows.sort(key=lambda r: (r[score_i] if isinstance(r[score_i], (int, float)) else -1), reverse=True)
        ws.delete_rows(2, ws.max_row)
        for r in rows:
            ws.append(r)

    wb.save(args.out)
    print(f"\n{done} screened, {needs_ocr} flagged NEEDS_OCR -> {args.out}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
