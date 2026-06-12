#!/usr/bin/env python3
"""PO → compliant label-instruction file (working demo of the make-or-break milestone).

Pipeline: read a purchase-order PDF → extract header + line items → EAN13 check-digit
validation → apply brand-licence rules + retailer/country rules from an EDITABLE rule
library (CSV here; Google Sheets/Airtable in production — same code, different reader)
→ write the label-instruction .xlsx with codes preserved as TEXT, plus a Review sheet.

Doctrine (matches the client's spec):
  * invalid EAN13s are BLOCKED from the output and flagged — no silent errors
  * a brand with no confirmed credit line is HELD and flagged — never guessed
  * retailer rules (hanger colour, alarm code, OEKO-TEX, label languages) come from the
    rule library, not from anyone's memory

Production notes (full build): extraction of varied/changing PDF layouts uses an LLM
(Claude) constrained to a strict JSON schema, followed by these same deterministic
validators — the LLM proposes, the validators dispose. This demo parses one layout
deterministically to prove the spine end-to-end without an API key.

Usage:  python3 po_pipeline.py sample_po_retailco.pdf --out label_instructions.xlsx
"""
from __future__ import annotations

import argparse
import csv
import re
import sys
from pathlib import Path

import pdfplumber
from openpyxl import Workbook


def ean13_valid(code: str) -> bool:
    if not re.fullmatch(r"\d{13}", code):
        return False
    d = [int(c) for c in code]
    chk = (10 - (sum(d[0:12:2]) + 3 * sum(d[1:12:2])) % 10) % 10
    return chk == d[12]


def load_rules(path: str) -> dict:
    out = {}
    with open(path, newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            key = row.get("brand") or row.get("retailer")
            out[key.strip()] = {k: (v or "").strip() for k, v in row.items()}
    return out


def extract(pdf_path: str) -> dict:
    """Extract header + line items from the PO. (Demo: deterministic for this layout;
    production: schema-constrained LLM extraction feeding the SAME validators.)"""
    with pdfplumber.open(pdf_path) as pdf:
        text = "\n".join(p.extract_text() or "" for p in pdf.pages)
        tables = [t for p in pdf.pages for t in (p.extract_tables() or [])]
    header = {
        "po_number": _find(r"PURCHASE ORDER\s+(\S+)", text),
        "retailer": _find(r"Retailer:\s*(.+?)\s+Brand licence:", text),
        "brand": _find(r"Brand licence:\s*(.+)", text),
        "country": _find(r"Country:\s*(\w+)", text),
    }
    items = []
    for table in tables:
        for row in table[1:]:                      # skip header row
            if row and row[0] and re.fullmatch(r"\d{13}", row[0].strip()):
                items.append({"ean13": row[0].strip(), "style": row[1], "size": row[2],
                              "composition": row[3], "qty": row[4], "unit_price": row[5]})
    return {"header": header, "items": items}


def _find(pat, text):
    m = re.search(pat, text)
    return m.group(1).strip() if m else ""


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("pdf")
    ap.add_argument("--brand-rules", default="brand_rules.csv")
    ap.add_argument("--retailer-rules", default="retailer_rules.csv")
    ap.add_argument("--out", default="label_instructions.xlsx")
    args = ap.parse_args()

    po = extract(args.pdf)
    brands = load_rules(args.brand_rules)
    retailers = load_rules(args.retailer_rules)
    h = po["header"]
    brand_rule = brands.get(h["brand"], {})
    retailer_rule = retailers.get(h["retailer"], {})

    flags = []
    if not brand_rule:
        flags.append(("HOLD", "brand", f"no rule entry for brand {h['brand']!r} — held, not guessed"))
    elif not brand_rule.get("credit_line"):
        flags.append(("HOLD", "brand", f"brand {h['brand']!r} has NO confirmed credit line — held, not guessed"))
    if not retailer_rule:
        flags.append(("HOLD", "retailer", f"no rule entry for retailer {h['retailer']!r}"))

    wb = Workbook()
    ws = wb.active
    ws.title = "Label Instructions"
    cols = ["EAN13", "Style", "Size", "Composition", "Qty", "Unit Price", "Credit Line",
            "Credit Year", "Label Languages", "Hanger Colour", "Alarm Code", "OEKO-TEX", "Status"]
    ws.append(cols)

    ok = blocked = 0
    for it in po["items"]:
        if not ean13_valid(it["ean13"]):
            flags.append(("BLOCKED", it["ean13"], "EAN13 check digit INVALID — excluded from output"))
            blocked += 1
            continue
        held = any(f[0] == "HOLD" for f in flags)
        row = [it["ean13"], it["style"], it["size"], it["composition"], it["qty"], it["unit_price"],
               brand_rule.get("credit_line", ""), brand_rule.get("year", ""),
               retailer_rule.get("label_languages", ""), retailer_rule.get("hanger_colour", ""),
               retailer_rule.get("alarm_code", ""),
               "yes" if retailer_rule.get("oeko_tex_required") == "yes" else "no",
               "HELD" if held else "OK"]
        ws.append(row)
        ws.cell(row=ws.max_row, column=1).number_format = "@"   # EAN preserved as TEXT
        ok += 1

    rv = wb.create_sheet("Review Flags")
    rv.append(["Severity", "Subject", "Detail"])
    for f in flags:
        rv.append(list(f))

    wb.save(args.out)
    print(f"PO {h['po_number']} | retailer={h['retailer']} | brand={h['brand']}")
    print(f"{ok} labels written, {blocked} BLOCKED (invalid EAN), {len(flags)} review flag(s) -> {args.out}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
